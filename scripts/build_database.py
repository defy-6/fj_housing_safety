"""Build the complete Fujian housing-safety SQLite analytical database."""

from __future__ import annotations

import hashlib
import json
import random
import sqlite3
from datetime import datetime, timezone
from pathlib import Path

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
SOURCE = ROOT / "data/raw/2024/housing-safety-export-2024-11-23.xlsx"
DATABASE = ROOT / "database/housing-safety.sqlite"
REPORT = ROOT / "database/database-report.md"
SCHEMA_FILE = ROOT / "database/schema.sql"

SHEETS = {
    "基础信息": {"key": "basic", "header_rows": 2, "data_row": 3, "topic": "房屋基础信息"},
    "安全隐患整治情况": {"key": "rectification", "header_rows": 4, "data_row": 5, "topic": "安全隐患整治"},
    "暂无安全隐患建筑": {"key": "potential_risk", "header_rows": 3, "data_row": 4, "topic": "暂无隐患建筑潜在风险"},
}

CITY_CODES = {
    "福州市": "350100", "厦门市": "350200", "莆田市": "350300", "三明市": "350400",
    "泉州市": "350500", "漳州市": "350600", "南平市": "350700", "龙岩市": "350800",
    "宁德市": "350900", "平潭综合试验区": "350128",
}
CITY_ALIASES = {"平潭综合实验区": "平潭综合试验区"}
TEXT_CORRECTIONS = {"钢筋混凝士结构": "钢筋混凝土结构"}


def canonical_city(name: str) -> str:
    return CITY_ALIASES.get(str(name).strip(), str(name).strip())


def clean_label(value: object) -> str:
    text = " ".join(str(value or "").replace("\n", " ").split())
    return TEXT_CORRECTIONS.get(text, text)


def merged_value_map(sheet, header_rows: int) -> dict[tuple[int, int], object]:
    values = {}
    for merged in sheet.merged_cells.ranges:
        if merged.min_row > header_rows:
            continue
        value = sheet.cell(merged.min_row, merged.min_col).value
        for row in range(merged.min_row, min(merged.max_row, header_rows) + 1):
            for col in range(merged.min_col, merged.max_col + 1):
                values[(row, col)] = value
    return values


def metric_definitions(sheet, config: dict) -> list[dict]:
    merged = merged_value_map(sheet, config["header_rows"])
    metrics = []
    for col in range(3, sheet.max_column + 1):
        path = []
        for row in range(1, config["header_rows"] + 1):
            label = clean_label(sheet.cell(row, col).value if sheet.cell(row, col).value is not None else merged.get((row, col)))
            if label and (not path or label != path[-1]):
                path.append(label)
        path_text = " / ".join(path) or f"未命名指标 {col}"
        metrics.append({
            "metric_id": f"{config['key']}.c{col:03d}",
            "sheet_name": sheet.title,
            "source_column": col,
            "metric_name": path[-1] if path else path_text,
            "metric_path": path_text,
            "topic": config["topic"],
            "category": path[0] if path else config["topic"],
        })
    return metrics


def region_key(city: str, county: str) -> tuple[str, str, str, str | None, str]:
    if city == "全省总计":
        return "350000", "福建省", "province", None, "standard"
    city = canonical_city(city)
    city_id = CITY_CODES[city]
    if county == "-":
        level = "special_city" if city == "平潭综合试验区" else "city"
        return city_id, city, level, "350000", "standard"
    digest = hashlib.sha1(f"{city}|{county}".encode()).hexdigest()[:10]
    return f"LOCAL-{digest}", county.strip(), "local", city_id, "source_unit"


SCHEMA = """
PRAGMA foreign_keys=ON;
CREATE TABLE metadata(key TEXT PRIMARY KEY, value TEXT NOT NULL);
CREATE TABLE import_batch(
  batch_id INTEGER PRIMARY KEY, source_file TEXT NOT NULL, source_sha256 TEXT NOT NULL,
  cutoff_date TEXT NOT NULL, imported_at TEXT NOT NULL, workbook_sheets INTEGER NOT NULL,
  status TEXT NOT NULL CHECK(status IN ('running','complete','failed'))
);
CREATE TABLE data_source(
  source_id INTEGER PRIMARY KEY, batch_id INTEGER NOT NULL REFERENCES import_batch(batch_id),
  sheet_name TEXT NOT NULL, header_rows INTEGER NOT NULL, first_data_row INTEGER NOT NULL,
  source_rows INTEGER NOT NULL, source_columns INTEGER NOT NULL, UNIQUE(batch_id,sheet_name)
);
CREATE TABLE dim_region(
  region_id TEXT PRIMARY KEY, region_name TEXT NOT NULL, region_level TEXT NOT NULL,
  parent_region_id TEXT REFERENCES dim_region(region_id), region_type TEXT NOT NULL,
  canonical_city TEXT, source_city_name TEXT, source_county_name TEXT
);
CREATE TABLE region_alias(
  alias_id INTEGER PRIMARY KEY, alias_name TEXT NOT NULL, region_id TEXT NOT NULL REFERENCES dim_region(region_id),
  alias_type TEXT NOT NULL, UNIQUE(alias_name,region_id)
);
CREATE TABLE dim_metric(
  metric_id TEXT PRIMARY KEY, sheet_name TEXT NOT NULL, source_column INTEGER NOT NULL,
  metric_name TEXT NOT NULL, metric_path TEXT NOT NULL, topic TEXT NOT NULL, category TEXT NOT NULL,
  unit TEXT NOT NULL DEFAULT 'count', value_type TEXT NOT NULL DEFAULT 'integer',
  aggregation_rule TEXT NOT NULL DEFAULT 'SUM', UNIQUE(sheet_name,source_column)
);
CREATE TABLE source_row_snapshot(
  snapshot_id INTEGER PRIMARY KEY, source_id INTEGER NOT NULL REFERENCES data_source(source_id),
  source_row INTEGER NOT NULL, region_id TEXT NOT NULL REFERENCES dim_region(region_id),
  source_city_name TEXT NOT NULL, source_county_name TEXT NOT NULL, row_json TEXT NOT NULL,
  UNIQUE(source_id,source_row)
);
CREATE TABLE simulation_run(
  run_id INTEGER PRIMARY KEY, scenario TEXT NOT NULL, base_year INTEGER NOT NULL,
  target_years TEXT NOT NULL, fluctuation_min REAL NOT NULL, fluctuation_max REAL NOT NULL,
  random_seed INTEGER NOT NULL, rule_version TEXT NOT NULL, created_at TEXT NOT NULL
);
CREATE TABLE fact_region_metric(
  fact_id INTEGER PRIMARY KEY, region_id TEXT NOT NULL REFERENCES dim_region(region_id),
  data_year INTEGER NOT NULL, metric_id TEXT NOT NULL REFERENCES dim_metric(metric_id),
  metric_value REAL NOT NULL, scenario TEXT NOT NULL, is_simulated INTEGER NOT NULL CHECK(is_simulated IN (0,1)),
  source_id INTEGER REFERENCES data_source(source_id), source_row INTEGER, source_column INTEGER,
  simulation_run_id INTEGER REFERENCES simulation_run(run_id), base_year INTEGER,
  UNIQUE(region_id,data_year,metric_id,scenario)
);
CREATE TABLE quality_check_result(
  check_id INTEGER PRIMARY KEY, batch_id INTEGER NOT NULL REFERENCES import_batch(batch_id),
  rule_code TEXT NOT NULL, severity TEXT NOT NULL CHECK(severity IN ('INFO','WARNING','ERROR')),
  status TEXT NOT NULL CHECK(status IN ('PASS','FAIL','OBSERVATION')),
  record_key TEXT, expected_value TEXT, actual_value TEXT, message TEXT NOT NULL,
  checked_at TEXT NOT NULL
);
CREATE INDEX idx_fact_year_metric ON fact_region_metric(data_year,metric_id,scenario);
CREATE INDEX idx_fact_region_year ON fact_region_metric(region_id,data_year,scenario);
CREATE INDEX idx_region_parent ON dim_region(parent_region_id);
CREATE INDEX idx_quality_rule ON quality_check_result(rule_code,status);
CREATE VIEW v_metric_values AS
SELECT f.data_year,f.scenario,f.is_simulated,r.region_id,r.region_name,r.region_level,r.parent_region_id,
       m.metric_id,m.metric_name,m.metric_path,m.topic,m.category,m.unit,f.metric_value
FROM fact_region_metric f JOIN dim_region r ON r.region_id=f.region_id JOIN dim_metric m ON m.metric_id=f.metric_id;
CREATE VIEW v_yearly_comparison AS
SELECT region_id,metric_id,
       MAX(CASE WHEN data_year=2024 THEN metric_value END) AS value_2024,
       MAX(CASE WHEN data_year=2025 THEN metric_value END) AS value_2025,
       MAX(CASE WHEN data_year=2026 THEN metric_value END) AS value_2026
FROM fact_region_metric GROUP BY region_id,metric_id;
"""


def add_check(connection, batch_id: int, rule: str, severity: str, status: str, message: str, expected=None, actual=None, key=None):
    connection.execute(
        "INSERT INTO quality_check_result(batch_id,rule_code,severity,status,record_key,expected_value,actual_value,message,checked_at) VALUES(?,?,?,?,?,?,?,?,?)",
        (batch_id, rule, severity, status, key, None if expected is None else str(expected), None if actual is None else str(actual), message, datetime.now(timezone.utc).isoformat()),
    )


def main() -> None:
    if DATABASE.exists():
        DATABASE.unlink()
    DATABASE.parent.mkdir(parents=True, exist_ok=True)
    SCHEMA_FILE.write_text(SCHEMA.strip() + "\n", encoding="utf-8")
    source_hash = hashlib.sha256(SOURCE.read_bytes()).hexdigest()
    workbook = load_workbook(SOURCE, data_only=True)
    connection = sqlite3.connect(DATABASE)
    connection.executescript(SCHEMA)
    now = datetime.now(timezone.utc).isoformat()
    connection.execute("INSERT INTO import_batch VALUES(1,?,?,?,?,?,?)", (SOURCE.name, source_hash, "2024-11-22", now, len(workbook.sheetnames), "running"))
    connection.executemany("INSERT INTO metadata VALUES(?,?)", [
        ("database_title", "福建省房屋安全完整分析数据库"), ("schema_version", "1.0.0"),
        ("actual_data_year", "2024"), ("simulation_years", "2025,2026"), ("simulation_range", "-5%,+5%"),
    ])

    connection.execute("INSERT INTO dim_region VALUES(?,?,?,?,?,?,?,?)", ("350000", "福建省", "province", None, "standard", None, "全省总计", "-"))
    for city, code in CITY_CODES.items():
        connection.execute("INSERT INTO dim_region VALUES(?,?,?,?,?,?,?,?)", (code, city, "special_city" if city == "平潭综合试验区" else "city", "350000", "standard", city, city, "-"))
    connection.execute("INSERT INTO region_alias(alias_name,region_id,alias_type) VALUES(?,?,?)", ("平潭综合实验区", "350128", "source_spelling"))

    source_ids = {}
    all_metrics = []
    local_regions = {}
    for source_id, (sheet_name, config) in enumerate(SHEETS.items(), 1):
        sheet = workbook[sheet_name]
        metrics = metric_definitions(sheet, config)
        all_metrics.extend(metrics)
        source_rows = sheet.max_row - config["data_row"] + 1
        source_ids[sheet_name] = source_id
        connection.execute("INSERT INTO data_source VALUES(?,?,?,?,?,?,?)", (source_id, 1, sheet_name, config["header_rows"], config["data_row"], source_rows, sheet.max_column))
        connection.executemany(
            "INSERT OR IGNORE INTO dim_metric(metric_id,sheet_name,source_column,metric_name,metric_path,topic,category) VALUES(:metric_id,:sheet_name,:source_column,:metric_name,:metric_path,:topic,:category)", metrics,
        )
        for row_number in range(config["data_row"], sheet.max_row + 1):
            row = [sheet.cell(row_number, col).value for col in range(1, sheet.max_column + 1)]
            city_raw, county_raw = str(row[0]).strip(), str(row[1]).strip()
            region_id, region_name, level, parent_id, region_type = region_key(city_raw, county_raw)
            if level == "local" and region_id not in local_regions:
                city = canonical_city(city_raw)
                connection.execute("INSERT INTO dim_region VALUES(?,?,?,?,?,?,?,?)", (region_id, region_name, level, parent_id, region_type, city, city_raw, county_raw))
                local_regions[region_id] = (city, county_raw)
            connection.execute(
                "INSERT INTO source_row_snapshot(source_id,source_row,region_id,source_city_name,source_county_name,row_json) VALUES(?,?,?,?,?,?)",
                (source_id, row_number, region_id, city_raw, county_raw, json.dumps(row, ensure_ascii=False, default=str, separators=(",", ":"))),
            )
            for metric in metrics:
                value = row[metric["source_column"] - 1]
                if value is None:
                    continue
                if not isinstance(value, (int, float)):
                    raise TypeError(f"Non-numeric metric at {sheet_name}!R{row_number}C{metric['source_column']}: {value!r}")
                connection.execute(
                    "INSERT INTO fact_region_metric(region_id,data_year,metric_id,metric_value,scenario,is_simulated,source_id,source_row,source_column) VALUES(?,?,?,?,?,?,?,?,?)",
                    (region_id, 2024, metric["metric_id"], float(value), "actual", 0, source_id, row_number, metric["source_column"]),
                )

    run_id = 1
    connection.execute("INSERT INTO simulation_run VALUES(?,?,?,?,?,?,?,?,?)", (run_id, "test_5pct", 2024, "[2025,2026]", -0.05, 0.05, 20242026, "1.0.0", now))
    rng = random.Random(20242026)
    detail_rows = connection.execute(
        "SELECT region_id,metric_id,metric_value FROM fact_region_metric WHERE data_year=2024 AND region_id LIKE 'LOCAL-%' ORDER BY region_id,metric_id"
    ).fetchall()
    prior = {(region_id, metric_id): value for region_id, metric_id, value in detail_rows}
    for year in (2025, 2026):
        current = {}
        for (region_id, metric_id), value in sorted(prior.items()):
            simulated = max(0, round(value * (1 + rng.uniform(-0.05, 0.05))))
            current[(region_id, metric_id)] = simulated
            connection.execute(
                "INSERT INTO fact_region_metric(region_id,data_year,metric_id,metric_value,scenario,is_simulated,simulation_run_id,base_year) VALUES(?,?,?,?,?,?,?,?)",
                (region_id, year, metric_id, simulated, "test_5pct", 1, run_id, year - 1),
            )
        prior = current
        connection.execute(
            """INSERT INTO fact_region_metric(region_id,data_year,metric_id,metric_value,scenario,is_simulated,simulation_run_id,base_year)
               SELECT r.parent_region_id,?,f.metric_id,SUM(f.metric_value),'test_5pct',1,?,?
               FROM fact_region_metric f JOIN dim_region r ON r.region_id=f.region_id
               WHERE f.data_year=? AND f.scenario='test_5pct' AND r.region_level='local'
               GROUP BY r.parent_region_id,f.metric_id""", (year, run_id, year - 1, year),
        )
        connection.execute(
            """INSERT INTO fact_region_metric(region_id,data_year,metric_id,metric_value,scenario,is_simulated,simulation_run_id,base_year)
               SELECT '350000',?,metric_id,SUM(metric_value),'test_5pct',1,?,?
               FROM fact_region_metric WHERE data_year=? AND scenario='test_5pct' AND region_id IN ({})
               GROUP BY metric_id""".format(",".join("?" for _ in CITY_CODES)),
            (year, run_id, year - 1, year, *CITY_CODES.values()),
        )

    expected_metrics = sum(workbook[name].max_column - 2 for name in SHEETS)
    metric_count = connection.execute("SELECT COUNT(*) FROM dim_metric").fetchone()[0]
    add_check(connection, 1, "METRIC_COUNT", "ERROR", "PASS" if metric_count == expected_metrics else "FAIL", "全部工作表指标均已建档", expected_metrics, metric_count)
    region_count = connection.execute("SELECT COUNT(*) FROM dim_region").fetchone()[0]
    add_check(connection, 1, "REGION_COUNT", "ERROR", "PASS" if region_count == 112 else "FAIL", "区域维表包含省、市和全部源明细单元", 112, region_count)
    for sheet_name, config in SHEETS.items():
        actual = connection.execute("SELECT source_rows FROM data_source WHERE sheet_name=?", (sheet_name,)).fetchone()[0]
        add_check(connection, 1, "SOURCE_ROW_COUNT", "ERROR", "PASS" if actual == 112 else "FAIL", f"{sheet_name}数据行数检查", 112, actual, sheet_name)
    negative = connection.execute("SELECT COUNT(*) FROM fact_region_metric WHERE metric_value<0").fetchone()[0]
    add_check(connection, 1, "NON_NEGATIVE", "ERROR", "PASS" if negative == 0 else "FAIL", "计数指标不得为负数", 0, negative)
    duplicates = connection.execute("SELECT COUNT(*) FROM (SELECT region_id,data_year,metric_id,scenario,COUNT(*) n FROM fact_region_metric GROUP BY 1,2,3,4 HAVING n>1)").fetchone()[0]
    add_check(connection, 1, "UNIQUE_FACT_KEY", "ERROR", "PASS" if duplicates == 0 else "FAIL", "事实表复合键唯一", 0, duplicates)
    for year in (2024, 2025, 2026):
        count = connection.execute("SELECT COUNT(*) FROM fact_region_metric WHERE data_year=?", (year,)).fetchone()[0]
        add_check(connection, 1, "YEAR_FACT_COUNT", "ERROR", "PASS" if count == 20160 else "FAIL", f"{year}年度事实记录完整", 20160, count, str(year))

    reconciliation = connection.execute(
        """SELECT s.region_id,s.metric_id,s.metric_value,COALESCE(d.detail_value,0),s.metric_value-COALESCE(d.detail_value,0)
           FROM fact_region_metric s LEFT JOIN (
             SELECT r.parent_region_id region_id,f.metric_id,SUM(f.metric_value) detail_value
             FROM fact_region_metric f JOIN dim_region r ON r.region_id=f.region_id
             WHERE f.data_year=2024 AND f.scenario='actual' AND r.region_level='local' GROUP BY 1,2
           ) d ON d.region_id=s.region_id AND d.metric_id=s.metric_id
           WHERE s.data_year=2024 AND s.scenario='actual' AND s.region_id IN ({}) AND ABS(s.metric_value-COALESCE(d.detail_value,0))>0.0001""".format(",".join("?" for _ in CITY_CODES)),
        tuple(CITY_CODES.values()),
    ).fetchall()
    for region_id, metric_id, source_value, detail_value, difference in reconciliation:
        add_check(connection, 1, "SOURCE_SUMMARY_RECONCILIATION", "WARNING", "OBSERVATION", "源地市汇总与源明细合计存在差异，已保留源汇总值", source_value, detail_value, f"{region_id}|{metric_id}|diff={difference}")
    add_check(connection, 1, "SIMULATION_REPRODUCIBLE", "INFO", "PASS", "模拟使用固定随机种子并由明细向上汇总", "seed=20242026", "seed=20242026")
    connection.execute("UPDATE import_batch SET status='complete' WHERE batch_id=1")
    connection.commit()

    stats = dict(connection.execute("SELECT 'regions',COUNT(*) FROM dim_region UNION ALL SELECT 'metrics',COUNT(*) FROM dim_metric UNION ALL SELECT 'facts',COUNT(*) FROM fact_region_metric UNION ALL SELECT 'snapshots',COUNT(*) FROM source_row_snapshot UNION ALL SELECT 'checks',COUNT(*) FROM quality_check_result"))
    failures = connection.execute("SELECT COUNT(*) FROM quality_check_result WHERE status='FAIL'").fetchone()[0]
    observations = connection.execute("SELECT COUNT(*) FROM quality_check_result WHERE status='OBSERVATION'").fetchone()[0]
    REPORT.write_text(f"""# 福建省房屋安全数据库构建报告

## 构建结果

- 数据库：`database/housing-safety.sqlite`
- 数据库版本：1.0.0
- 原始年份：2024
- 模拟年份：2025、2026（逐年 ±5%，固定随机种子 20242026）
- 区域：{stats['regions']} 个
- 指标：{stats['metrics']} 个
- 事实记录：{stats['facts']:,} 条
- 原始行快照：{stats['snapshots']} 条
- 质量检查：{stats['checks']} 条
- 失败检查：{failures} 条
- 汇总差异观察项：{observations} 条

## 核心表

| 表/视图 | 用途 |
|---|---|
| `dim_region` | 省、市、区县/开发区/片区区域维表 |
| `dim_metric` | 180 个完整指标及原始列位置 |
| `fact_region_metric` | 2024 实际与 2025/2026 模拟事实长表 |
| `source_row_snapshot` | 三张源表每一行的完整 JSON 快照 |
| `simulation_run` | 模拟场景、浮动范围和随机种子 |
| `quality_check_result` | 完整性和勾稽检查结果 |
| `v_metric_values` | 已关联区域和指标名称的查询视图 |
| `v_yearly_comparison` | 2024—2026 横向比较视图 |

## 查询约定

- 实际数据：`scenario='actual' AND is_simulated=0`
- 测试数据：`scenario='test_5pct' AND is_simulated=1`
- 每条 2024 事实均保存源工作表、源行、源列。
- 质量观察项不改写原始数据；源汇总与明细差异均保存在 `quality_check_result`。
""", encoding="utf-8")
    connection.close()
    print(json.dumps({**stats, "failures": failures, "observations": observations, "database": str(DATABASE)}, ensure_ascii=False))


if __name__ == "__main__":
    main()
