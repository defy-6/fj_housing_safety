"""Build browser-ready housing safety data and Fujian map assets."""

from __future__ import annotations

import json
import random
import sqlite3
import struct
from pathlib import Path

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
SOURCE = ROOT / "data/raw/2024/housing-safety-export-2024-11-23.xlsx"
MAP_DIR = ROOT / "data/reference/administrative/geojson/cities"
PUBLIC = ROOT / "apps/web/public/data"
GPKG = ROOT / "data/raw/fujian-geo/fujian.gpkg"
DATABASE = ROOT / "database/housing-safety.sqlite"

CITY_NAMES = ["福州市", "厦门市", "莆田市", "三明市", "泉州市", "漳州市", "南平市", "龙岩市", "宁德市", "平潭综合试验区"]
ALIASES = {"平潭综合实验区": "平潭综合试验区"}


def normalized_city(value: str) -> str:
    return ALIASES.get(value, value)


def metric_payload(row: tuple) -> dict:
    total = int(row[25] or 0)
    major = int(row[2] or 0)
    general = int(row[3] or 0)
    no_hazard = int(row[4] or 0)
    hazard = major + general
    return {
        "inspectionTotal": total,
        "majorHazard": major,
        "generalHazard": general,
        "noHazard": no_hazard,
        "hazardRate": round(hazard / max(1, major + general + no_hazard), 6),
        "age": {
            "60年代及以前": int(row[6] or 0),
            "70年代": int(row[7] or 0),
            "80年代": int(row[8] or 0),
            "90年代": int(row[9] or 0),
            "2000年代": int(row[10] or 0),
            "2010年代及以后": int(row[11] or 0),
        },
        "structure": {
            "钢结构": int(row[13] or 0),
            "框架结构": int(row[14] or 0),
            "砖混结构": int(row[15] or 0),
            "砌体结构": int(row[20] or 0),
            "钢筋混凝土结构": int(row[22] or 0),
            "其他": int(row[24] or 0),
        },
    }


def fluctuate(payload: dict, rng: random.Random) -> dict:
    def n(value: int) -> int:
        return max(0, round(value * (1 + rng.uniform(-0.05, 0.05))))

    result = {
        "inspectionTotal": n(payload["inspectionTotal"]),
        "majorHazard": n(payload["majorHazard"]),
        "generalHazard": n(payload["generalHazard"]),
        "noHazard": n(payload["noHazard"]),
        "age": {key: n(value) for key, value in payload["age"].items()},
        "structure": {key: n(value) for key, value in payload["structure"].items()},
    }
    hazard = result["majorHazard"] + result["generalHazard"]
    result["hazardRate"] = round(hazard / max(1, hazard + result["noHazard"]), 6)
    return result


def build_dashboard() -> dict:
    workbook = load_workbook(SOURCE, read_only=True, data_only=True)
    sheet = workbook["基础信息"]
    city_actual: dict[str, dict] = {}
    county_actual: dict[str, dict] = {}
    for row in sheet.iter_rows(min_row=3, values_only=True):
        city = normalized_city(row[0]) if row[0] else ""
        county = row[1]
        if not city or city == "全省总计":
            continue
        payload = metric_payload(row)
        if county == "-":
            city_actual[city] = payload
        elif county:
            county_actual[f"{city}|{county}"] = {"city": city, "county": county, **payload}

    years: dict[str, dict] = {"2024": {"cities": city_actual, "counties": county_actual}}
    rng = random.Random(20242026)
    previous_cities = city_actual
    previous_counties = county_actual
    for year in (2025, 2026):
        cities = {name: fluctuate(value, rng) for name, value in previous_cities.items()}
        counties = {
            key: {"city": value["city"], "county": value["county"], **fluctuate(value, rng)}
            for key, value in previous_counties.items()
        }
        years[str(year)] = {"cities": cities, "counties": counties}
        previous_cities, previous_counties = cities, counties

    return {
        "meta": {
            "title": "福建省房屋安全动态监测",
            "actualCutoff": "2024-11-22",
            "simulationRange": "±5%",
            "randomSeed": 20242026,
            "cityCount": len(city_actual),
            "countyRecordCount": len(county_actual),
        },
        "years": years,
    }


def build_analytics() -> dict:
    """Export all 180 database metrics and three annual slices for browser analysis."""
    with sqlite3.connect(DATABASE) as connection:
        connection.row_factory = sqlite3.Row
        metrics = [dict(row) for row in connection.execute(
            "SELECT metric_id,metric_name,metric_path,topic,category,unit FROM dim_metric ORDER BY topic,category,source_column"
        )]
        regions = [dict(row) for row in connection.execute(
            "SELECT region_id,region_name,region_level,parent_region_id,canonical_city,region_type FROM dim_region ORDER BY region_level,canonical_city,region_name"
        )]
        values: dict[str, dict[str, dict[str, int | float]]] = {}
        for row in connection.execute(
            "SELECT data_year,region_id,metric_id,metric_value FROM fact_region_metric ORDER BY data_year,region_id,metric_id"
        ):
            value = row["metric_value"]
            normalized = int(value) if value == int(value) else round(value, 4)
            values.setdefault(str(row["data_year"]), {}).setdefault(row["region_id"], {})[row["metric_id"]] = normalized
        checks = dict(connection.execute(
            "SELECT status,COUNT(*) FROM quality_check_result GROUP BY status"
        ).fetchall())
    return {
        "meta": {"metricCount": len(metrics), "regionCount": len(regions), "years": [2024, 2025, 2026], "qualityChecks": checks},
        "metrics": metrics,
        "regions": regions,
        "values": values,
    }


def simplify_line(points: list[list[float]], tolerance: float) -> list[list[float]]:
    if len(points) <= 2:
        return points
    keep = {0, len(points) - 1}
    stack = [(0, len(points) - 1)]
    tolerance_sq = tolerance * tolerance
    while stack:
        start, end = stack.pop()
        ax, ay = points[start]
        bx, by = points[end]
        dx, dy = bx - ax, by - ay
        length_sq = dx * dx + dy * dy
        best_index, best_distance = -1, 0.0
        for index in range(start + 1, end):
            px, py = points[index]
            if length_sq:
                t = max(0.0, min(1.0, ((px - ax) * dx + (py - ay) * dy) / length_sq))
                qx, qy = ax + t * dx, ay + t * dy
                distance = (px - qx) ** 2 + (py - qy) ** 2
            else:
                distance = (px - ax) ** 2 + (py - ay) ** 2
            if distance > best_distance:
                best_index, best_distance = index, distance
        if best_index >= 0 and best_distance > tolerance_sq:
            keep.add(best_index)
            stack.extend([(start, best_index), (best_index, end)])
    return [points[index] for index in sorted(keep)]


def simplify_ring(ring: list[list[float]], tolerance: float) -> list[list[float]]:
    points = ring[:-1] if len(ring) > 1 and ring[0] == ring[-1] else ring
    if len(points) <= 4:
        return points + [points[0]]
    origin = points[0]
    split = max(range(1, len(points)), key=lambda index: (points[index][0] - origin[0]) ** 2 + (points[index][1] - origin[1]) ** 2)
    first = simplify_line(points[:split + 1], tolerance)
    second = simplify_line(points[split:] + [origin], tolerance)
    result = first[:-1] + second[:-1]
    return result + [result[0]] if len(result) >= 3 else ring


def read_wkb(data: bytes, offset: int = 0) -> tuple[dict, int]:
    byte_order = data[offset]
    endian = "<" if byte_order == 1 else ">"
    raw_type = struct.unpack_from(f"{endian}I", data, offset + 1)[0]
    offset += 5
    dimensions = 2
    geometry_type = raw_type
    if raw_type >= 3000:
        dimensions, geometry_type = 4, raw_type - 3000
    elif raw_type >= 2000:
        dimensions, geometry_type = 3, raw_type - 2000
    elif raw_type >= 1000:
        dimensions, geometry_type = 3, raw_type - 1000

    def point() -> list[float]:
        nonlocal offset
        values = struct.unpack_from(f"{endian}{'d' * dimensions}", data, offset)
        offset += 8 * dimensions
        return [round(values[0], 7), round(values[1], 7)]

    if geometry_type == 1:
        return {"type": "Point", "coordinates": point()}, offset
    if geometry_type == 3:
        ring_count = struct.unpack_from(f"{endian}I", data, offset)[0]
        offset += 4
        rings = []
        for _ in range(ring_count):
            count = struct.unpack_from(f"{endian}I", data, offset)[0]
            offset += 4
            rings.append([point() for _ in range(count)])
        return {"type": "Polygon", "coordinates": rings}, offset
    if geometry_type == 6:
        count = struct.unpack_from(f"{endian}I", data, offset)[0]
        offset += 4
        polygons = []
        for _ in range(count):
            geometry, offset = read_wkb(data, offset)
            polygons.append(geometry["coordinates"])
        return {"type": "MultiPolygon", "coordinates": polygons}, offset
    raise ValueError(f"Unsupported WKB geometry type: {raw_type}")


def decode_gpkg_geometry(blob: bytes, tolerance: float = 0.00018) -> dict:
    flags = blob[3]
    envelope_code = (flags >> 1) & 0b111
    envelope_values = {0: 0, 1: 4, 2: 6, 3: 6, 4: 8}[envelope_code]
    geometry, _ = read_wkb(blob, 8 + envelope_values * 8)
    if geometry["type"] == "Polygon":
        geometry["coordinates"] = [simplify_ring(ring, tolerance) for ring in geometry["coordinates"]]
    elif geometry["type"] == "MultiPolygon":
        geometry["coordinates"] = [[simplify_ring(ring, tolerance) for ring in polygon] for polygon in geometry["coordinates"]]
    return geometry


def layer_features(table: str, include_city: bool = False) -> dict:
    city_by_prefix = {"3501": "福州市", "3502": "厦门市", "3503": "莆田市", "3504": "三明市", "3505": "泉州市", "3506": "漳州市", "3507": "南平市", "3508": "龙岩市", "3509": "宁德市"}
    features = []
    with sqlite3.connect(GPKG) as connection:
        for code, name, blob in connection.execute(f'SELECT XZQDM, XZQMC, geom FROM "{table}" ORDER BY fid'):
            canonical_name = "平潭综合试验区" if "平潭综合实" in name else name.strip()
            properties = {"name": canonical_name, "adcode": str(code)}
            if include_city:
                properties["city"] = "平潭综合试验区" if str(code) == "350128" else city_by_prefix[str(code)[:4]]
            features.append({"type": "Feature", "properties": properties, "geometry": decode_gpkg_geometry(blob)})
    return {"type": "FeatureCollection", "features": features}


def main() -> None:
    (PUBLIC / "maps/cities").mkdir(parents=True, exist_ok=True)
    (PUBLIC / "housing-safety.json").write_text(json.dumps(build_dashboard(), ensure_ascii=False, separators=(",", ":")))
    (PUBLIC / "housing-analytics.json").write_text(json.dumps(build_analytics(), ensure_ascii=False, separators=(",", ":")))
    (PUBLIC / "maps/fujian-province.json").write_text(json.dumps(layer_features("省调查界线"), ensure_ascii=False, separators=(",", ":")))
    (PUBLIC / "maps/fujian-regions.json").write_text(json.dumps(layer_features("设区市调查界线"), ensure_ascii=False, separators=(",", ":")))
    (PUBLIC / "maps/fujian-counties.json").write_text(json.dumps(layer_features("县级调查界线", include_city=True), ensure_ascii=False, separators=(",", ":")))
    for source in MAP_DIR.glob("*.json"):
        (PUBLIC / "maps/cities" / source.name).write_bytes(source.read_bytes())


if __name__ == "__main__":
    main()
